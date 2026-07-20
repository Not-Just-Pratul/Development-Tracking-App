"""
Report generation module for PDF and Excel exports
"""
from flask import Blueprint, send_file, abort, make_response
from io import BytesIO
from datetime import datetime

from models import Project, Phase, Stage, Step, User
from secure_auth import login_required

reports = Blueprint('reports', __name__, url_prefix='/reports')


@reports.get('/project/<int:project_id>.pdf')
@login_required
def export_project_pdf(project_id):
    """Export project details as PDF"""
    try:
        project = Project.query.get_or_404(project_id)
    except Exception as e:
        from flask import flash, redirect, url_for
        flash(f'Error loading project: {str(e)}', 'error')
        return redirect(url_for('ui.projects_list'))
    
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError as e:
        # If reportlab is not installed, return a detailed error
        from flask import flash, redirect, url_for
        flash(f'PDF export requires reportlab library. Error: {str(e)}. Please run: pip install reportlab', 'error')
        return redirect(url_for('ui.project_detail', project_id=project_id))
    
    # Create PDF buffer
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # Container for PDF elements
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#3498db'),
        spaceAfter=12,
        spaceBefore=12
    )
    
    # Title
    elements.append(Paragraph(f"Project Report: {project.name}", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Project Information
    project_info = [
        ['Project Name:', project.name],
        ['Project Head:', project.head.full_name if project.head else 'Not Assigned'],
        ['Head Designation:', project.head_designation.designation_name if project.head_designation else '-'],
        ['Head Department:', project.head_department.department_name if project.head_department else '-'],
        ['Part Name:', project.part_name or '-'],
        ['Part Code:', project.part_code or '-'],
        ['Customer Name:', project.customer_name or '-'],
        ['Company Name:', project.company_name or '-'],
        ['Location:', project.location_rel.location_name if project.location_rel else '-'],
        ['Start Date:', project.start_date.strftime('%d %b %Y') if project.start_date else '-'],
        ['Expected End:', project.expected_end_date.strftime('%d %b %Y') if project.expected_end_date else '-'],
        ['Progress:', f"{project.progress_percent}%"],
        ['Remarks:', (project.part_description[:60] + '...') if project.part_description and len(project.part_description) > 60 else (project.part_description or '-')],
        ['Generated:', datetime.now().strftime('%d %b %Y %H:%M')]
    ]
    
    info_table = Table(project_info, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#ecf0f1')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey)
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Phases and Stages
    for phase in project.phases:
        elements.append(Paragraph(f"Phase {phase.serial_number}: {phase.name}", heading_style))
        
        if phase.stages:
            stage_data = [['S.No', 'Stage Name', 'Department', 'Responsible', 'Status']]
            
            for stage in phase.stages:
                status = 'Completed' if stage.actual_end_date else 'In Progress'
                department = stage.responsible_department.department_name if stage.responsible_department else '-'
                responsible = stage.responsible.full_name if stage.responsible else 'Unassigned'
                
                stage_data.append([
                    str(stage.serial_number),
                    stage.name[:35] + '...' if len(stage.name) > 35 else stage.name,
                    department[:15] + '...' if len(department) > 15 else department,
                    responsible[:20] + '...' if len(responsible) > 20 else responsible,
                    status
                ])
            
            stage_table = Table(stage_data, colWidths=[0.4*inch, 2.2*inch, 1.2*inch, 1.3*inch, 0.9*inch])
            stage_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
            ]))
            
            elements.append(stage_table)
            elements.append(Spacer(1, 0.2*inch))
        else:
            elements.append(Paragraph("No stages defined", styles['Normal']))
            elements.append(Spacer(1, 0.2*inch))
    
    # Build PDF
    try:
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'project_{project.id}_{project.name.replace(" ", "_")}.pdf'
        )
    except Exception as e:
        from flask import flash, redirect, url_for
        flash(f'Error generating PDF: {str(e)}', 'error')
        return redirect(url_for('ui.project_detail', project_id=project_id))


@reports.get('/project/<int:project_id>.xlsx')
@login_required
def export_project_excel(project_id):
    """Export project details as Excel"""
    try:
        project = Project.query.get_or_404(project_id)
    except Exception as e:
        from flask import flash, redirect, url_for
        flash(f'Error loading project: {str(e)}', 'error')
        return redirect(url_for('ui.projects_list'))
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        # If openpyxl is not installed, return a detailed error
        from flask import flash, redirect, url_for
        flash(f'Excel export requires openpyxl library. Error: {str(e)}. Please run: pip install openpyxl', 'error')
        return redirect(url_for('ui.project_detail', project_id=project_id))
    
    # Create workbook
    wb = Workbook()
    
    # Project Overview Sheet
    ws_overview = wb.active
    assert ws_overview is not None  # Workbook always has an active sheet
    ws_overview.title = "Project Overview"
    
    # Header styling
    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Title
    ws_overview['A1'] = f"Project Report: {project.name}"
    ws_overview['A1'].font = Font(bold=True, size=16, color="2c3e50")
    ws_overview.merge_cells('A1:E1')
    
    # Project Information
    ws_overview['A3'] = "Project Information"
    ws_overview['A3'].font = Font(bold=True, size=14)
    
    info_data = [
        ['Project Name:', project.name],
        ['Project Head:', project.head.full_name if project.head else 'Not Assigned'],
        ['Head Designation:', project.head_designation.designation_name if project.head_designation else '-'],
        ['Head Department:', project.head_department.department_name if project.head_department else '-'],
        ['Part Name:', project.part_name or '-'],
        ['Part Code:', project.part_code or '-'],
        ['Customer Name:', project.customer_name or '-'],
        ['Company Name:', project.company_name or '-'],
        ['Location:', project.location_rel.location_name if project.location_rel else '-'],
        ['Start Date:', project.start_date.strftime('%d %b %Y') if project.start_date else '-'],
        ['Expected End:', project.expected_end_date.strftime('%d %b %Y') if project.expected_end_date else '-'],
        ['Overall Progress:', f"{project.progress_percent}%"],
        ['Total Phases:', len(project.phases)],
        ['Remarks:', project.part_description or '-'],
        ['Generated:', datetime.now().strftime('%d %b %Y %H:%M')]
    ]
    
    row = 5
    for label, value in info_data:
        ws_overview[f'A{row}'] = label
        ws_overview[f'A{row}'].font = Font(bold=True)
        ws_overview[f'B{row}'] = value
        row += 1
    
    # Adjust column widths
    ws_overview.column_dimensions['A'].width = 20
    ws_overview.column_dimensions['B'].width = 40
    
    # Stages Detail Sheet
    ws_stages = wb.create_sheet("Stages Detail")
    
    # Headers
    headers = ['Phase', 'S.No', 'Stage Name', 'Department', 'Responsible Person', 'Start Date', 
               'Planned End', 'Actual End', 'Progress %', 'Status']
    
    for col, header in enumerate(headers, 1):
        cell = ws_stages.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    row = 2
    for phase in project.phases:
        for stage in phase.stages:
            status = 'Completed' if stage.actual_end_date else 'In Progress'
            department = stage.responsible_department.department_name if stage.responsible_department else '-'
            responsible = stage.responsible.full_name if stage.responsible else 'Unassigned'
            
            ws_stages.cell(row=row, column=1).value = f"{phase.serial_number}. {phase.name}"
            ws_stages.cell(row=row, column=2).value = stage.serial_number
            ws_stages.cell(row=row, column=3).value = stage.name
            ws_stages.cell(row=row, column=4).value = department
            ws_stages.cell(row=row, column=5).value = responsible
            ws_stages.cell(row=row, column=6).value = stage.start_date.strftime('%d/%m/%Y') if stage.start_date else '-'
            ws_stages.cell(row=row, column=7).value = stage.expected_end_date.strftime('%d/%m/%Y') if stage.expected_end_date else '-'
            ws_stages.cell(row=row, column=8).value = stage.actual_end_date.strftime('%d/%m/%Y') if stage.actual_end_date else '-'
            ws_stages.cell(row=row, column=9).value = f"{stage.progress_percent}%"
            ws_stages.cell(row=row, column=10).value = status
            
            # Alternate row colors
            if row % 2 == 0:
                fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
                for col in range(1, 11):
                    ws_stages.cell(row=row, column=col).fill = fill
            
            row += 1
    
    # Adjust column widths
    for col in range(1, 11):
        ws_stages.column_dimensions[get_column_letter(col)].width = 18
    ws_stages.column_dimensions['C'].width = 35  # Stage name column wider
    
    # Save to buffer
    try:
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'project_{project.id}_{project.name.replace(" ", "_")}.xlsx'
        )
    except Exception as e:
        from flask import flash, redirect, url_for
        flash(f'Error generating Excel file: {str(e)}', 'error')
        return redirect(url_for('ui.project_detail', project_id=project_id))
